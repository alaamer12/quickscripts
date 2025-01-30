from openpyxl import load_workbook

# Load the Excel file
file_path = r"C:\Users\amrmu\Downloads\Backlog.xlsx"
output_file_path = r"C:\Users\amrmu\Downloads\Backlog Sorted.xlsx"
sheet_name = 'Sheet1'  # Replace with your actual sheet name if necessary

# Load the workbook and the sheet
workbook = load_workbook(filename=file_path)
sheet = workbook[sheet_name]

# Extract all rows of data, including the header
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Sort the data based on the values in the 5th column (index 4)
header = data[0]  # Save the header row
data = data[1:]   # Data without header
sorted_data = sorted(data, key=lambda x: x[4] if x[4] is not None else '')

# Write the sorted data back into the sheet
for i, row in enumerate([header] + sorted_data, start=1):
    for j, value in enumerate(row, start=1):
        sheet.cell(row=i, column=j, value=value)

# Save the updated workbook
workbook.save(filename=output_file_path)
