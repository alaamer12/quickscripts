from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Border, Side, Alignment, Font,
    NamedStyle
)
from openpyxl.utils import get_column_letter


def enhance_table_styling(
        filepath: str,
        table_start_cell: str = 'A1',
        table_end_column: str = 'I',
        has_header: bool = True,
        min_row_height: float = 30,  # Minimum row height in points
        vertical_padding: float = 5  # Padding in points
) -> None:
    """
    Enhanced table styling with proper vertical spacing, content-based row heights,
    and status-based color coding for column E.

    Args:
        filepath: Path to the Excel file
        table_start_cell: Starting cell of the table (e.g., 'A1')
        table_end_column: Last column of the table (e.g., 'I')
        has_header: Whether the table has a header row
        min_row_height: Minimum height for rows in points
        vertical_padding: Additional vertical padding in points
    """
    # Define status color mapping
    color_map = {
        'Completed': ('C6EFCE', '006100'),  # Light green bg, Dark green text
        'Pending': ('FFEB9C', '9C6500'),  # Light yellow bg, Dark brown text
        'Repeatedly': ('FFC7CE', '9C0006'),  # Light red bg, Dark red text
        'This Weak': ('BDD7EE', '1F497D'),  # Light blue bg, Dark blue text
        'Today': ('FFEB9C', '9C6500'),  # Light yellow bg, Dark brown text
        'Working on it': ('FFC7CE', '9C0006'),  # Light red bg, Dark red text
    }

    wb = load_workbook(filepath)
    ws = wb.active

    # Determine table dimensions
    start_col = table_start_cell[0]
    start_row = int(table_start_cell[1:])
    end_row = ws.max_row

    # Define styles
    header_font = Font(
        name='Arial',
        size=11,
        bold=True,
        color='FFFFFF'
    )

    cell_font = Font(
        name='Arial',
        size=10
    )

    header_fill = PatternFill(
        start_color='366092',
        end_color='366092',
        fill_type='solid'
    )

    # Define borders
    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define alignment with proper vertical spacing
    wrapped_alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=False,
        indent=1
    )

    # Set column widths (basic autofit)
    for col in range(ord(start_col) - 64, ord(table_end_column) - 64 + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        column = ws[column_letter]

        for cell in column:
            try:
                if cell.value:
                    # Calculate length considering newlines
                    lines = str(cell.value).split('\n')
                    max_length = max(max_length, max(len(line) for line in lines))
            except:
                pass

        # Add padding to width and set minimum width
        adjusted_width = max(max_length + 2, 10)  # Reduced padding for better fit
        ws.column_dimensions[column_letter].width = adjusted_width

    # First pass: Apply styles and basic formatting
    for row in range(start_row, end_row + 1):
        for col in range(ord(start_col) - 64, ord(table_end_column) - 64 + 1):
            cell = ws[f'{get_column_letter(col)}{row}']

            # Apply alignment and wrap text first
            cell.alignment = wrapped_alignment

            # Apply header or regular cell styles
            if row == start_row and has_header:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = medium_border
            else:
                # Check if this is column E (status column) and apply color mapping
                if get_column_letter(col) == 'E' and cell.value:
                    status = str(cell.value).strip()
                    if status in color_map:
                        bg_color, text_color = color_map[status]
                        cell.fill = PatternFill(
                            start_color=bg_color,
                            end_color=bg_color,
                            fill_type='solid'
                        )
                        cell.font = Font(
                            name='Arial',
                            size=10,
                            color=text_color
                        )
                else:
                    cell.font = cell_font
                    if row % 2 == 0:
                        cell.fill = PatternFill(
                            start_color='F2F2F2',
                            end_color='F2F2F2',
                            fill_type='solid'
                        )
                cell.border = thin_border

    # Second pass: Calculate and set row heights
    for row in range(start_row, end_row + 1):
        max_row_height = 0
        for col in range(ord(start_col) - 64, ord(table_end_column) - 64 + 1):
            cell = ws[f'{get_column_letter(col)}{row}']
            if cell.value:
                # Calculate height based on content and column width
                content = str(cell.value)
                column_width = ws.column_dimensions[get_column_letter(col)].width

                # Calculate wrapped lines
                chars_per_line = int(column_width * 1.8)  # Characters that fit in column
                lines = []
                current_line = []
                words = content.split()

                current_length = 0
                for word in words:
                    word_length = len(word)
                    if current_length + word_length + 1 <= chars_per_line:
                        current_line.append(word)
                        current_length += word_length + 1
                    else:
                        lines.append(' '.join(current_line))
                        current_line = [word]
                        current_length = word_length

                if current_line:
                    lines.append(' '.join(current_line))

                # Calculate height needed (15 points per line is typical)
                content_height = len(lines) * 15
                max_row_height = max(max_row_height, content_height)

        # Add padding and ensure minimum height
        final_height = max(min_row_height, max_row_height + (vertical_padding * 2))
        ws.row_dimensions[row].height = final_height

    # Add table borders
    # Add table borders
    for row in range(start_row, end_row + 1):
        # Left border of first column
        ws[f'{start_col}{row}'].border = Border(
            left=Side(style='medium'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Right border of last column
        ws[f'{table_end_column}{row}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='medium'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    wb.save(filepath)

    # Example usage:
if __name__ == "__main__":
    enhance_table_styling(
        filepath=r"C:\Users\amrmu\Downloads\Task list 2025.xlsx",
        table_start_cell='A1',
        table_end_column='I',
        has_header=True,
        min_row_height=30,
        vertical_padding=5
    )
