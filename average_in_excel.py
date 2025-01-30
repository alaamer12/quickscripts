import openpyxl

def is_countable(value):
    """Check if the value is countable (integer or float)."""
    return isinstance(value, (int, float))

def calculate_average_for_sheet(sheet):
    """Calculate the average of countable values in column A from A3 to the last row in the given sheet."""
    total = 0
    count = 0
    for row in sheet.iter_rows(min_row=3, min_col=1, max_col=1):
        cell = row[0]
        if cell.value is not None and is_countable(cell.value):
            total += cell.value
            count += 1
    return total / count if count > 0 else 0

def calculate_averages(file_path):
    """Calculate averages for all sheets and return the results."""
    workbook = openpyxl.load_workbook(file_path)
    sheet_averages = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        average = calculate_average_for_sheet(sheet)
        sheet_averages[sheet_name] = average
    return sheet_averages

def calculate_overall_average(sheet_averages):
    """Calculate the overall average of the averages from all sheets, excluding zeros."""
    total = 0
    count = 0
    for average in sheet_averages.values():
        if average != 0:
            total += average
            count += 1
    return total / count if count > 0 else 0

def main():
    file_path = r"C:\Users\amrmu\Downloads\نتيجة 2 عام للاعلان.xlsx"
    sheet_averages = calculate_averages(file_path)
    for sheet_name, average in sheet_averages.items():
        print(f"The average value from A3 to the last row of column A in sheet '{sheet_name}' is: {average}")

    overall_average = calculate_overall_average(sheet_averages)
    print(f"The overall average of the averages from all sheets (excluding zeros) is: {overall_average}")

if __name__ == "__main__":
    main()
    x = input("Press")